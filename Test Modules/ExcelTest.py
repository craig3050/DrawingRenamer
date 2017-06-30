import os
import openpyxl
from openpyxl import Workbook

def excel_list(file_path):
    drawing_list = Workbook()
    worksheet_1 = drawing_list.active
    drawing_list.save("Drawing_list.xlsx")
    row_number = 1
    for file_name in os.listdir(file_path):
        worksheet_1.cell(row = row_number, column = 1).value = file_name
        row_number += 1
    drawing_list.save("Drawing_list.xlsx")
    #break before writing the new numbers
    continue_on()

    #rename the files
    for i in range (1, row_number):
        new_file_name = worksheet_1.cell(row = i, column = 2).value
        old_file_name = worksheet_1.cell(row = i, column = 1).value
        print ("Old = " + old_file_name + "New = " + new_file_name)



#option to continue or break programme
def continue_on():
    continue_onwards = input("Do you wish to continue? enter 'Y' or 'y': ")
    if continue_onwards == "Y" or continue_onwards == "y":
        return
    else:
        exit(0)







if __name__ == '__main__':
    print("Welcome to Craig's simple file renaming tool!\n\n")
    file_path = input("Enter the Path of the file to be processed: ")
    excel_list(file_path)
