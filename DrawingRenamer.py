import PyPDF2
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook

#need pip install pypdf2, pip install openpyxl
#if error with UTF / charmap encoding run the following command in the CMD line - chcp 65001

#option to continue or break programme
def continue_on():
    continue_onwards = input("Do you wish to continue? enter 'Y' or 'y': ")
    if continue_onwards == "Y" or continue_onwards == "y":
        return
    else:
        exit(0)

def extract_pdf_text(file_path):
    with open(file_path, 'rb') as file_open:
        file_object = PyPDF2.PdfFileReader(file_open)
        page_object = file_object.getPage(0)
        page_object = page_object.extractText()
        #print (page_object) #comment this out for general use - useful when testing if the drawing text has been recognised
    return page_object


def RenameFile(file_path, new_drawing_number):
    source = file_path.split("/")[-1]
    file_extension = source.split(".")[-1]
    destination = file_path.replace(source, new_drawing_number + "." + file_extension)
    os.rename(file_path, destination)


def RegExTest(input_text, regex_format):
    output_text = ""
    regex_variable = "."
    for char in regex_format:
        if char == "-":
            output_text += "-"
        elif char == "_":
            output_text += "_"
        elif char == "(":
            output_text += "\("
        elif char == ")":
            output_text += "\)"
        else:
            output_text += regex_variable
    regex_results = re.findall(output_text, input_text)
    number_of_regex_groups = len(regex_results) -1
    return regex_results[number_of_regex_groups]

def sanitisedFilePath(file_path):
    sanitised_file_path = ""
    for char in file_path:
        if char == "\"":
            sanitised_file_path += ""
        elif char == "\\":
            sanitised_file_path += "/"
        else:
            sanitised_file_path += char
    return sanitised_file_path


def excel_list(file_path):
    drawing_list = Workbook()
    worksheet_1 = drawing_list.active
    drawing_list.save("Drawing_list.xlsx")
    row_number = 1
    for file_name in os.listdir(file_path):
        file_name_no_extension = ""
        for char in file_name:
            if char != ".":
                file_name_no_extension += char
            else:
                break
        file_extension = file_name.split(".")[-1]
        worksheet_1.cell(row = row_number, column = 1).value = file_name_no_extension
        worksheet_1.cell(row = row_number, column = 2).value = file_extension
        row_number += 1
    drawing_list.save("Drawing_list.xlsx")
    print("The drawing list has been written to the same directory as the script")
    print("Add your drawings into column C")
    #break before writing the new numbers
    continue_on()

    drawing_list = load_workbook("Drawing_list.xlsx")
    worksheet_1 = drawing_list.active
    #rename the files
    for i in range (1, row_number):
        old_file_name = worksheet_1.cell(row = i, column = 1).value
        print(old_file_name)
        file_extension = worksheet_1.cell(row=i, column = 2).value
        print(file_extension)
        new_file_name = worksheet_1.cell(row = i, column = 3).value
        print(new_file_name)
        print("Old = " + old_file_name + ",  New = " + new_file_name)
        old_file_name = file_path + "/" + old_file_name + "." + file_extension
        RenameFile(old_file_name, new_file_name)


def select_option_1(file_path):
    #Text remove / replace function in filename
    #Use this if lots of the drawing number is the same
    text_remove = input("Enter the text you would like to remove / replace: ")
    text_replace = input("Enter the text you would like to add. Enter nothing for delete only: ")
    continue_on()
    for file_name in os.listdir(file_path):
        source = sanitisedFilePath(file_path)
        source += ("/" + file_name)
        if text_remove in file_name:
            destination = file_name.replace(text_remove, text_replace)
            destination = destination.split(".")[0]
            RenameFile(source, destination)


def select_option_2(file_path):
    #add text to the beginning of the filename
    text_add = input("Enter the text you would like to add to the beginning of the file name: ")
    continue_on()
    for file_name in os.listdir(file_path):
        source = sanitisedFilePath(file_path)
        source += ("/" + file_name)
        destination = text_add + file_name
        destination = destination.split(".")[0]
        RenameFile(source, destination)



def select_option_3(file_path):
    #add text to the end of the filename
    text_add = input("Enter the text you would like to add to the end of the file name: ")
    continue_on()
    for file_name in os.listdir(file_path):
        source = sanitisedFilePath(file_path)
        source += ("/" + file_name)
        destination = file_name.split(".")[0]
        destination = destination + text_add
        RenameFile(source, destination)



def select_option_4(file_path):
    #rename the file based on what is in the titleblock
    regex_template = input("Enter a sample drawing number: ")
    continue_on()
    for file_name in os.listdir(file_path):
        print(file_name)
        sanitised_file_path = sanitisedFilePath(file_path)
        sanitised_file_path += "/" + file_name
        pdf_text = extract_pdf_text(sanitised_file_path)
        pdf_text_match = RegExTest(pdf_text, regex_template)
        RenameFile(sanitised_file_path, pdf_text_match)



def select_option_5(file_path):
    excel_list(sanitisedFilePath(file_path))




if __name__ == '__main__':
    print("Welcome to Craig's simple file renaming tool!\n\n")
    file_path = input("Enter the Path of the folder to be processed: ")
    print("Printing directory contents: \n")
    for file_name in os.listdir(file_path):
        print(file_name)
        print("\n")
    print (""" \n\n
    What would you like to do: \n
    1. Text Remove / Replace Function
    2. Add Text to Beginning of File Name
    3. Add Text to End of File Name
    4. Rename drawings to whatever is in the PDF titleblock
    5. Export drawings to an excel list, then re-name files based on the second column
    """)
    choice = input("Please enter your choice: ")
    if choice == "1":
        select_option_1(file_path)
    elif choice == "2":
        select_option_2(file_path)
    elif choice == "3":
        select_option_3(file_path)
    elif choice == "4":
        select_option_4(file_path)
    elif choice == "5":
        select_option_5(file_path)
    else:
        print ("Not a valid choice\nBye bye")

